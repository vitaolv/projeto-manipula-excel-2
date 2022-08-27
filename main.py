import openpyxl

import classes

i = 3  # count for rows
i2 = 0  # count for list
iAB = 2  # count for A's and B's rows
iBR = 3  # count for blue's and red's rows

try:
    acess_date = classes.import_data(pach='./Resources/Ranked_lol')

    id = int(input('Choose a game match id: (Available: 1 to 17126)\n'))

    if 0 < id < 17126:
        blue, red, wins = acess_date.select(id)
        create = classes.work()

    else:
        raise ValueError(f'You entered number "{id}" out of range.')

    t = acess_date.get_time()
    print(t)

    new_date = create.add('Dados')

    create.add_text(match_time=t, wins=wins)

    for row in acess_date.action:
        a = str(row)
        b = str(blue[i2])
        r = str(red[i2])

        create.get_processing_header(cell=f'A{i}', action=a)
        create.get_processing_int(cell=f'C{i}', action=b)
        create.get_processing_int(cell=f'D{i}', action=r)
        i = i + 1
        i2 = i2 + 1

    create.merge(cell_init='A1', cell_end='D1')
    create.apply_styles(
        cel='A1',
        styles=[
            ('font', openpyxl.styles.Font(b=True, sz=16, color='ded4a2')),
            ('fill', openpyxl.styles.PatternFill('solid', fgColor='59606e')),
            ('alignment', openpyxl.styles.Alignment(
                vertical="center", horizontal='center')),
            ('border', openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                              right=openpyxl.styles.Side(
                                                  style='thin'),
                                              top=openpyxl.styles.Side(
                                                  style='thin'),
                                              bottom=openpyxl.styles.Side(style='thin'))
             )
        ]
    )

    for rows in range(13):
        create.merge(cell_init=f'A{iAB}', cell_end=f'B{iAB}')

        create.apply_styles(
            cel=f'A{iAB}',
            styles=[
                ('font', openpyxl.styles.Font(b=True, sz=10, color='ffffff')),
                ('fill', openpyxl.styles.PatternFill('solid', fgColor='59606e')),
                ('alignment', openpyxl.styles.Alignment(
                    vertical="center", horizontal='center')),
                ('border', openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                  right=openpyxl.styles.Side(
                                                      style='thin'),
                                                  top=openpyxl.styles.Side(
                                                      style='thin'),
                                                  bottom=openpyxl.styles.Side(style='thin'))
                 )
            ]
        )

        create.apply_styles(
            cel=f'B{iAB}',
            styles=[
                ('font', openpyxl.styles.Font(b=True, sz=10, color='ffffff')),
                ('fill', openpyxl.styles.PatternFill('solid', fgColor='59606e')),
                ('alignment', openpyxl.styles.Alignment(
                    vertical="center", horizontal='center')),
                ('border', openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                  right=openpyxl.styles.Side(
                                                      style='thin'),
                                                  top=openpyxl.styles.Side(
                                                      style='thin'),
                                                  bottom=openpyxl.styles.Side(style='thin'))

                 )
            ]
        )

        iAB = iAB + 1

    create.apply_styles(
        cel='C2',
        styles=[
            ('font', openpyxl.styles.Font(b=True, sz=12, color='ded4a2')),
            ('fill', openpyxl.styles.PatternFill('solid', fgColor='1531bd')),
            ('alignment', openpyxl.styles.Alignment(
                vertical="center", horizontal='center')),
            ('border', openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                              right=openpyxl.styles.Side(
                                                  style='thin'),
                                              top=openpyxl.styles.Side(
                                                  style='thin'),
                                              bottom=openpyxl.styles.Side(style='thin'))
             )
        ]
    )

    create.apply_styles(
        cel='D2',
        styles=[
            ('font', openpyxl.styles.Font(b=True, sz=12, color='ded4a2')),
            ('fill', openpyxl.styles.PatternFill('solid', fgColor='7d0000')),
            ('alignment', openpyxl.styles.Alignment(
                vertical="center", horizontal='center')),
            ('border', openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                              right=openpyxl.styles.Side(
                                                  style='thin'),
                                              top=openpyxl.styles.Side(
                                                  style='thin'),
                                              bottom=openpyxl.styles.Side(style='thin'))
             )
        ]
    )

    for rows in range(12):
        create.apply_styles(
            cel=f'D{iBR}',
            styles=[
                ('font', openpyxl.styles.Font(b=True, sz=10, color='ffffff')),
                ('fill', openpyxl.styles.PatternFill('solid', fgColor='bf5858')),
                ('alignment', openpyxl.styles.Alignment(
                    vertical="center", horizontal='center')),
                ('border', openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                  right=openpyxl.styles.Side(
                                                      style='thin'),
                                                  top=openpyxl.styles.Side(
                                                      style='thin'),
                                                  bottom=openpyxl.styles.Side(style='thin'))
                 )
            ]
        )

        create.apply_styles(
            cel=f'C{iBR}',
            styles=[
                ('font', openpyxl.styles.Font(b=True, sz=12, color='ffffff')),
                ('fill', openpyxl.styles.PatternFill('solid', fgColor='5593ad')),
                ('alignment', openpyxl.styles.Alignment(
                    vertical="center", horizontal='center')),
                ('border', openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                  right=openpyxl.styles.Side(
                                                      style='thin'),
                                                  top=openpyxl.styles.Side(
                                                      style='thin'),
                                                  bottom=openpyxl.styles.Side(style='thin'))
                 )
            ]
        )

        iBR = iBR + 1

    create.merge(cell_init='A15', cell_end='D16')
    create.apply_styles(
        cel='A15',
        styles=[
            ('font', openpyxl.styles.Font(b=True, sz=18, color='ded4a2')),
            ('fill', openpyxl.styles.PatternFill('solid', fgColor='59606e')),
            ('alignment', openpyxl.styles.Alignment(
                vertical="center", horizontal='center')),
            ('border', openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                              right=openpyxl.styles.Side(
                                                  style='thin'),
                                              top=openpyxl.styles.Side(
                                                  style='thin'),
                                              bottom=openpyxl.styles.Side(style='thin'))
             )
        ]
    )

    # So let's reference and crate graphs.
    cat = openpyxl.chart.Reference(new_date, min_col=1, min_row=3, max_row=14)
    info = openpyxl.chart.Reference(
        new_date, min_col=3, max_col=4, min_row=2, max_row=14)

    create.add_graph(celula='E1', categories=cat, dates=info)

    create.merge(cell_init='A17', cell_end='D25')
    create.add_img(celula='A17', path='./Resources/im.png')

    create.save(path='./Out/out.xlsx')

except FileNotFoundError as er:
    print(f'Error! File nor found.\nDetali: {er}')

except ValueError as err:
    print(f'Error! Enter within the range 1 to 17126.\nDetali: {err}')


if __name__ == '__main__':
    pass
