import time

import openpyxl


class import_data:
    def __init__(self, pach: str):
        self.pach = pach
        self.datasForTime = []  # type: list[str]
        self.count = 0
        self.blue = []          # type: list[str]
        self.red = []           # type: list[str]

        self.wins = ''

        self.action = ['KILLS', 'DEATH', 'ASSIST', 'DAMAGE DEALT', 'TOTAL HEAL', 'WARD PLACED', 'WARD KILLS',
                       'TOWER KILLS',
                       'TOTAL DRAGON KILLS', 'TOTAL BARON KILLS', 'TOTAL MINION KILLS', 'TOTAL GOLD']

    def select(self, array: int):
        with open(f'{self.pach}.csv', 'r') as arq:
            datas = arq.readlines()
            self.datasForTime = datas[array].replace('\n', '').split(',')

        # kills
        self.blue.append(self.datasForTime[14])
        self.red.append(self.datasForTime[38])

        # death
        self.blue.append(self.datasForTime[15])
        self.red.append(self.datasForTime[39])

        # assist
        self.blue.append(self.datasForTime[16])
        self.red.append(self.datasForTime[40])

        # damage
        self.blue.append(self.datasForTime[17])
        self.red.append(self.datasForTime[41])

        # heal
        self.blue.append(self.datasForTime[24])
        self.red.append(self.datasForTime[48])

        # ward
        self.blue.append(self.datasForTime[12])
        self.red.append(self.datasForTime[36])

        # ward kills
        self.blue.append(self.datasForTime[13])
        self.red.append(self.datasForTime[37])

        # tower kills
        self.blue.append(self.datasForTime[10])
        self.red.append(self.datasForTime[34])

        # dragon
        self.blue.append(self.datasForTime[8])
        self.red.append(self.datasForTime[32])

        # baron
        self.blue.append(self.datasForTime[9])
        self.red.append(self.datasForTime[33])

        # minions
        self.blue.append(self.datasForTime[19])
        self.red.append(self.datasForTime[43])

        # gold total
        self.blue.append(self.datasForTime[18])
        self.red.append(self.datasForTime[42])

        # time
        self.count = int(self.datasForTime[1])

        # wins
        if self.datasForTime[2] == '1':
            self.wins = 'BLUE'
        elif self.datasForTime[26] == '1':
            self.wins = 'RED'
        else:
            self.wins = 'UNDEFINED'

        return self.blue, self.red, self.wins

    def get_time(self):
        return time.strftime("%H:%M:%S", time.gmtime(self.count))


class work:
    def __init__(self):
        self.work = openpyxl.Workbook()
        self.datas = None

    def active(self):
        self.datas = work.active

    def add(self, title: str = ''):
        new = self.work.create_sheet(title)
        self.datas = new

        return new

    def add_text(self, match_time: int, wins: str):
        self.datas.title = 'MATCH HISTORY'
        self.datas['A1'] = 'MATCH HISTORY'
        self.datas['A2'] = f'Time: {match_time}'
        self.datas['C2'] = 'BLUE'
        self.datas['D2'] = 'RED'
        self.datas['A15'] = f'{wins} WINS!'

    # store str type in cell
    def get_processing_header(self, cell, action):
        self.datas[cell] = action

    # store int type in cell
    def get_processing_int(self, cell, action):
        self.datas[cell].value = int(action)

    def merge(self, cell_init: str, cell_end: str):
        self.datas.merge_cells(f'{cell_init}:{cell_end}')

    def apply_styles(self, cel: str, styles):
        for style in styles:
            setattr(self.datas[cel], style[0], style[1])

    def add_graph(self, celula, categories: openpyxl.chart.Reference, dates: openpyxl.chart.Reference):
        graph = openpyxl.chart.BarChart()
        sz = openpyxl.drawing.text.CharacterProperties(sz=200)
        graph.grouping = "percentStacked"
        graph.type = 'bar'
        graph.overlap = 100

        # graph's size
        graph.height = 14
        graph.width = 24

        graph.title = 'Match history graph'
        graph.y_axis.title = "Percent"

        graph.add_data(dates, titles_from_data=True)
        graph.set_categories(categories)

        graph.dataLabels = openpyxl.chart.label.DataLabelList()
        graph.dataLabels.showVal = True
        graph.dataLabels.showVal = openpyxl.chart.text.RichText(
            p=[openpyxl.chart.text.Paragraph(pPr=openpyxl.drawing.text.ParagraphProperties(defRPr=sz), endParaRPr=sz)])

        self.datas.add_chart(graph, celula)

    def add_img(self, celula: str, path: str):
        img = openpyxl.drawing.image.Image(path)
        self.datas.add_image(img, celula)

    def save(self, path: str):
        self.work.save(path)
